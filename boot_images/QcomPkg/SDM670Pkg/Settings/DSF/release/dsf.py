"""
  Description
    This file contains a DSF class which is used for DCB generation and ConfigMon
  ---------------------------------------------------------------------------
  Copyright (c) 2017 Qualcomm Technologies, Inc.
  All Rights Reserved. Qualcomm Technologies Proprietary and Confidential.
  ---------------------------------------------------------------------------
"""

# GLOBAL_USES_UNDERSCORE_AND_UPPERCASE
# local_uses_underscore_and_lowercase
# membersUseCamelCase
# FunctionsUseCamelCaseStartingWithCapital

import time
def status_message_and_time(string, function, *args):
    print("[STATUS] {:s} - Start".format(string))
    start = time.time()
    retval = function(*args)
    end = time.time()
    print("[STATUS] {:s} - End".format(string))
    print("[STATUS] Elapsed time {:d}ms".format(int((end - start) * 1000)))
    return retval

print("[STATUS] Import modules - Start")
start = time.time()
import re
import struct
import sys
import glob
import os
import stat
try:
    import xlrd
    import openpyxl as px
    import t32api
except ImportError:
    pass
end = time.time()
print("[STATUS] Import modules - End")
print("[STATUS] Elapse time {:d}ms".format(int((end - start) * 1000)))

DEBUG = False
# single = re.compile(r'^([A-Z_0-9]+)([a-z])([A-Z_0-9]*)[_]*<([0-9])>$')
# double = re.compile(r'^([A-Z_0-9]+)([a-z])([A-Z_0-9]*)([a-z])[_]*([A-Z_0-9]*)<([0-9])><([0-9])>$')
single = re.compile(r'^([A-Z_0-9]+)([a-z])([A-Z_0-9]*)<([0-9]+)>$')
double = re.compile(r'^([A-Z_0-9]+)([a-z])([A-Z_0-9]*)([a-z])([A-Z_0-9]*)_<([0-9]+)><([0-9]+)>$')

class DSF:
    # one-time class variables
    binaries_ordered_by_header_index = None
    globals = None
    dsf_xl_columns = ['Block Name','Relative Address','Register','Recommended Settings','Default','Compare flag','Notes','PoR Source','Comment','Clock Switch Hang','Traffic Stall','Trained']
    def __init__(self, filename=None, target_folder=None):
        """
        Operates on DSF XL file.

        Arguments:
        filename    --  name of DSF XL file
        """
        print("[INFO] Creating DSF object for {:s}".format(filename))
        # this class relies of relative paths
        # need to cd into file dir for relative paths to match
        os.chdir(os.path.dirname(os.path.abspath(__file__)))
        self.filename = filename
        if filename is not None:
            import xlrd
            self.dsfWorkbook = xlrd.open_workbook(self.filename)
        
        else:
            self.dsfWorkbook = None
        
        self.autoDict = {}
        self.absDict = {}
        self.relDict = {}
        self.sheetDictcprefixs = {} # check
        self.dcbBinDict = {}
        self.regDict = {}
        self.xlfileDict = {}
        self.trainingSheetDict = {}
        self.trainingRegDict = {}
        self.baseDict = {}
        self.scorecardDict = {}
        self.freqDict ={}
        self.configmonDict = {}
        self.absDict_from_T32 = {} # check
        self.abs_Add_value_dict = {} # check
        # set up one-time class variables if they have not been set yet
        if DSF.globals is None:
            DSF.create_globals(target_folder)
        if DSF.binaries_ordered_by_header_index is None:
            DSF.binaries_order_elements_function()
    
    @staticmethod
    def bytes_crc32(bytes):
        CRC_POLYNOMIAL = 0x04C11DB7
        CRC_INITVAL = 0xFFFFFFFF
        CRC_XOROUT = 0xFFFFFFFF

        result = CRC_INITVAL
        for octet in bytes:
            if type(octet) is str:
                octet = struct.unpack("B", octet)[0]
            for bit in range(0,8):
                if (octet >> 7) ^ (result >> 31):
                    result = ((result << 1) & 0xFFFFFFFF) ^ CRC_POLYNOMIAL
                else:
                    result = (result << 1) & 0xFFFFFFFF
                octet = (octet << 1) & 0xFF

        result = result ^ CRC_XOROUT
        return result
    
    @staticmethod
    def binaries_order_elements_function():
        #print(DSF.globals["FILES"])
        with open(DSF.globals["FILES"]["dcb.h"],"r") as dcb_header_fp:
            struct_flag=0
            DSF.binaries_ordered_by_header_index = []
            for line in dcb_header_fp:
                line=((re.sub('[\(\)\{\}<>]', '  ',line)).strip("#\n\t")).split(' ')#.split("., ")
                line=list(filter(None, line))
                if line !=[] and line[0] == 'enum':
                    struct_flag = 1
                elif line !=[] and struct_flag >=5 and line[0].find("DCB_IDX_MAX")==-1:
                    header_text = re.sub('_IDX','',line[0])
                    header_text = re.sub(',','',header_text)
                    header_text = re.sub('_CFG','',header_text)
                    DSF.binaries_ordered_by_header_index.append(header_text)
                    struct_flag += 1
                elif line !=[] and line[0].find("DCB_IDX_MAX")!=-1:
                    break
                elif struct_flag >0:
                    struct_flag += 1

    @staticmethod
    def rreplace(s, old, new, occurrence):
        li = s.rsplit(old, occurrence)
        return new.join(li)
    @staticmethod
    def create_globals(target_folder):
        DSF.globals = {}
        DSF.globals["PERFORCE_PATHS"] = True
        DSF.globals["NUM_CH"] = 0
        DSF.globals["NUM_DQ_PCH"] = 0
        DSF.globals["NUM_CA_PCH"] = 0
        DSF.globals["NUM_DQ_TRAINING_STACK"] = 0
        DSF.globals["NUM_CA_TRAINING_STACK"] = 0
        DSF.globals["TARGET_DDR_SYSTEM_FIRMWARE_MAJOR_VERSION"] = 0
        DSF.globals["TARGET_DDR_SYSTEM_FIRMWARE_MINOR_VERSION"] = 0
        DSF.globals["DCB_MAX_SIZE"] = 0

        sw_globals_files = {
            "ddrss_training.h": "../../../Library/DSFTargetLib/common/ddrss/header/ddrss_training.h",
            "ddr_phy_technology": "../../../Library/DSFTargetLib/common/target/" + target_folder + "/header/ddr_phy_technology.h",
            "dcb.h": "../dcb.h",
            "target_config.h": "../target_config.h",
        }

        sve_globals_files = {
            "ddrss_training.h": "../../common/ddrss/header/ddrss_training.h",
            "ddr_phy_technology": "../../common/target/" + target_folder + "/header/ddr_phy_technology.h",
            "dcb.h": "../../ddrsns/target/" + target_folder + "/settings/dcb.h",
            "target_config.h": "../../ddrsns/target/" + target_folder + "/settings/target_config.h",
        }
        # check which file paths to use
        if os.path.isfile(sve_globals_files["dcb.h"]):
            DSF.globals["FILES"] = sve_globals_files
            DSF.globals["COMMON"] = "../../common/target/" + target_folder + "/header/"
            DSF.globals["release"] = "../../ddrsns/target/" + target_folder + "/settings/release/"
            DSF.globals["target"] = "target/" + target_folder + "/"
            DSF.globals["external"] = DSF.globals["target"] + "/external/"
        else:
            DSF.globals["PERFORCE_PATHS"] = False
            DSF.globals["FILES"] = sw_globals_files
            DSF.globals["COMMON"] = "../../../Library/DSFTargetLib/common/target/" + target_folder + "/header/"
            DSF.globals["release"] = "../release/"
            DSF.globals["target"] = "../tools/common/target/" + target_folder + "/"
            DSF.globals["external"] = "../release/"
        #print(DSF.globals["FILES"]) 
        """
        Search for the following in sequence:
            #define
            one or more spaces and/or tabs
            macro name in all caps with underscores between words
            one or more spaces and/or tabs
            formula consisting of ()+-*/ and numbers
        """
        define_re = re.compile(r'^#define[ \t]+([A-Z_]+)[ \t]+([\(\)\+\-\*\/0-9 ]+)$')
        
        for filename in DSF.globals["FILES"]:
            with open(DSF.globals["FILES"][filename], "r") as fp:
                for line in fp:
                    # Remove comments
                    line = line.split("//")[0]
                    
                    # Check for a macro defined as a value or formula
                    m = define_re.match(line)
                    if m:
                        if DEBUG:
                            print("{:s} = {:s}".format(m.group(1), m.group(2)))
                        
                        # Check if macro name matches any key in globals dictionary
                        for key in DSF.globals:
                            if m.group(1) == key:
                                DSF.globals[key] = eval(m.group(2))
        
        if DEBUG:
            print(DSF.globals)

    @staticmethod
    def ConvertIndexRegisterName(reg):
        # Check for reg with one lower case character to number replacements
        m = single.match(reg)
        if m:
            reg_single = m.group(1)+m.group(4)+m.group(3)
            if reg_single[-1] == "_":
                reg_single = DSF.rreplace(reg_single, '_', '', 1)
            return reg_single
        # Check for reg with two lower case character to number replacements
        m = double.match(reg)
        if m:
            reg_double = m.group(1)+m.group(6)+m.group(3)+m.group(7)+m.group(5)
            if reg_double[-1] == "_":
                reg_double = DSF.rreplace(reg_double, '_', '', 1)
            return reg_double

        return reg
    def PopulateAutoDict(self, force=False):
        """
        Populates autoDict from AUTO sheet in DSF XL file.
        """
        if len(self.autoDict) != 0 and force == False:
            return

        # Regular expressions used to prefixes with lowercase character to number replacements
        single = re.compile(r'^([A-Z_0-9]+)([a-z])([A-Z_0-9]*)[_]*<([0-9])>$')
        double = re.compile(r'^([A-Z_0-9]+)([a-z])([A-Z_0-9]*)([a-z])[_]*([A-Z_0-9]*)<([0-9])><([0-9])>$')

        # Open AUTO sheet in DSF XL file
        autoSheet = self.dsfWorkbook.sheet_by_name('AUTO')

        # Read header from AUTO sheet
        headerArray = autoSheet.row_values(0)

        # Obtain column index for each header
        colIdxDict = {}
        for colIdx in range(len(headerArray)):
            colIdxDict[headerArray[colIdx]] = colIdx

        # Read data from AUTO sheet
        for rowIdx in range(1, autoSheet.nrows):
            rowArray = autoSheet.row_values(rowIdx)

            # Create dictionary for block name
            blockName = rowArray[colIdxDict['Block Name']]
            self.autoDict[blockName] = {}

            # Populate dictionary for block name
            self.autoDict[blockName]['DCB Binary Sub-Block'] = rowArray[colIdxDict['DCB Binary Sub-Block']]
            self.autoDict[blockName]['SWC Sheet'] = rowArray[colIdxDict['SWC Sheet']]
            self.autoDict[blockName]['C Code Register Offset'] = rowArray[colIdxDict['C Code Register Offset']]
            self.autoDict[blockName]['C Code HWIO Prefix']= rowArray[colIdxDict['C Code HWIO Prefix']]
            self.autoDict[blockName]['Summary Prefix']= rowArray[colIdxDict['Summary Prefix']]
            self.autoDict[blockName]['Remove Prefix']= int(rowArray[colIdxDict['Remove Prefix']])
            self.autoDict[blockName]['Header'] = rowArray[colIdxDict['Header']]
            if 'XL file' in colIdxDict:
                self.autoDict[blockName]['XL file'] = rowArray[colIdxDict['XL file']]
            if 'Summary Sheet' in colIdxDict:
                self.autoDict[blockName]['Summary Sheet'] = rowArray[colIdxDict['Summary Sheet']].split(',')
            self.autoDict[blockName]['Scorecard Index'] = int(rowArray[colIdxDict['Scorecard Index']])
            if 'Link' in colIdxDict:
                self.autoDict[blockName]['Link'] = rowArray[colIdxDict['Link']]

            # adding the xlfile as key and intializing its value to 0
            if 'XL file' in colIdxDict:
                self.xlfileDict[self.autoDict[blockName]['XL file']] = 0
            
            # updating and adding sheet Dictionary to store Cprefixes coressponding to it
            if rowArray[colIdxDict['SWC Sheet']] not in self.sheetDictcprefixs:
                self.sheetDictcprefixs[rowArray[colIdxDict['SWC Sheet']]] = []
            self.sheetDictcprefixs[rowArray[colIdxDict['SWC Sheet']]] = self.sheetDictcprefixs[rowArray[colIdxDict['SWC Sheet']]] + rowArray[colIdxDict['C Code HWIO Prefix']].split(',')

            # Perform lowercase character to number replacements if needed
            for prefix in rowArray[colIdxDict['CMM HWIO Prefixes']].split(','):
                self.autoDict[blockName]['CMM HWIO Prefixes'] = []

                # Check for prefix with one lowercase character to number replacement
                m = single.match(prefix)
                if m:
                    for i in range(int(m.group(4))+1):
                        self.autoDict[blockName]['CMM HWIO Prefixes'].append(\
                            m.group(1)+str(i)+m.group(3))
                    continue

                # Check for prefix with two lower case character to number replacements
                m = double.match(prefix)
                if m:
                    for i in range(int(m.group(6))+1):
                        for j in range(int(m.group(7))+1):
                            self.autoDict[blockName]['CMM HWIO Prefixes'].append(\
                                m.group(1)+str(i)+m.group(3)+str(j)+m.group(5))
                    continue

                # Prefix does not have one or two number replacements
                self.autoDict[blockName]['CMM HWIO Prefixes'].append(prefix)

    def PopulateDcbBinDictFromFiles(self):
        # Read binary files into dictionary
        for bin in DSF.binaries_ordered_by_header_index:
            # Only populate binaries which are not already present
            if bin not in self.dcbBinDict:
                name = bin + ".bin"
                with open(DSF.globals["target"] + "/bins/" + name, "rb") as f:
                    self.dcbBinDict[bin] = f.read()

    def PopulateDcbBinDictFromExt(self, dcb, binnames):
        # populate the dict from the existing dcb, override with binnames
        with open(dcb, "rb") as dcb_f:
            checksum = struct.unpack("<I", dcb_f.read(4))[0]
            dcb_size = struct.unpack("<I", dcb_f.read(4))[0]
            version = struct.unpack("<I", dcb_f.read(4))[0]

            header_offset = 12
            for bin in DSF.binaries_ordered_by_header_index:
                bin_off = struct.unpack("<H", dcb_f.read(2))[0]
                bin_size = struct.unpack("<H", dcb_f.read(2))[0]
                header_offset += 4

                print("bin check: "+bin)
                if bin in binnames:
                    with open(binnames[bin], "rb") as bin_f:
                        self.dcbBinDict[bin] = bin_f.read()
                else:
                    dcb_f.seek(bin_off)
                    self.dcbBinDict[bin] = dcb_f.read(bin_size)
                    dcb_f.seek(header_offset)
                    
    def CustomerDcbBinary(self, filename, populate_first=True, output=None):
        if output is None:
            output = DSF.globals["release"]
            
        self.PopulateDcbBinDictFromExt(os.path.join(output, filename), [])
        self.PopulateDcbBinDictFromBinarySheets()
        self.PopulateRegDictFromSecondarySheets()
        
        # Find and replace values
        dcb_blocks = {}
        reg_blocks_sorted = sorted(self.regDict.keys(), key=lambda obj: obj)
        for reg_block in reg_blocks_sorted:
            reg_names_sorted = sorted(self.regDict[reg_block].keys(), key=lambda obj: self.regDict[reg_block][obj]['Relative Address'])

            for regName in reg_names_sorted:
                dcb_block = self.regDict[reg_block][regName]['DCB Binary Sub-Block']
                
                if dcb_block not in dcb_blocks:
                    dcb_blocks[dcb_block] = {}
                    dcb_blocks[dcb_block]['Addresses'] = []
                    dcb_blocks[dcb_block]['Values'] = []
                    ba = bytearray(self.dcbBinDict[dcb_block])
                    while len(ba):
                        dcb_blocks[dcb_block]['Values'].append((ba.pop() << 24) + (ba.pop() << 16) + (ba.pop() << 8) + ba.pop())
                        dcb_blocks[dcb_block]['Addresses'].append((ba.pop() << 24) + (ba.pop() << 16) + (ba.pop() << 8) + ba.pop())
                
                # Get index from address
                address = self.regDict[reg_block][regName]['Relative Address']
                index = dcb_blocks[dcb_block]['Addresses'].index(address)
                
                # Replace value using index
                recommended = self.regDict[reg_block][regName]['Recommended Settings']
                dcb_blocks[dcb_block]['Values'][index] = recommended
                
        for dcb_block in dcb_blocks:
            self.dcbBinDict[dcb_block] = bytes()
            while len(dcb_blocks[dcb_block]['Addresses']):
                self.dcbBinDict[dcb_block] += struct.pack("<I", dcb_blocks[dcb_block]['Addresses'].pop())
                self.dcbBinDict[dcb_block] += struct.pack("<I", dcb_blocks[dcb_block]['Values'].pop())
        
        self.WriteDcbBinary(filename, output)
    
    def GenerateDcbBinary(self, filename, populate_first=True, output=None):
        if populate_first:
            self.PopulateDcbBinDictFromSheets()
            self.PopulateDcbBinDictFromFiles()
        
        self.WriteDcbBinary(filename, output)
        
    def WriteDcbBinary(self, filename, output=None):
        # Print information on which XL file is being processed and resulting binary
        print("[INFO] DSF XL {:s} to DCB binary {:s}".format(self.filename, filename))
        
        header_max_size = 100

        # Header bytes object
        header = bytes()

        # Data bytes object
        data = bytes()

        # Data starts at end of header
        offset = header_max_size

        # Loop through each binary section of DCB
        for bin in DSF.binaries_ordered_by_header_index:
            # Update header
            size = len(self.dcbBinDict[bin])
            header += struct.pack("<H", offset)
            header += struct.pack("<H", size)
            offset += size

            # Update data
            data += self.dcbBinDict[bin]
            
            # Align each section of DCB to 4 byte boundary
            remainder = offset % 4
            for _ in range(remainder):
                data += struct.pack("<B", 0)
            offset += remainder

        # DSF Version
        header = struct.pack("<H", DSF.globals["TARGET_DDR_SYSTEM_FIRMWARE_MAJOR_VERSION"]) + header
        header = struct.pack("<H", DSF.globals["TARGET_DDR_SYSTEM_FIRMWARE_MINOR_VERSION"]) + header

        # Size of DCB including header
        size = len(data) + header_max_size
        header = struct.pack("<I", size) + header

        # CRC set to zero as placeholder for CRC calculation
        header = struct.pack("<I",0) + header

        # Pad header with zeros
        header = header.ljust(header_max_size, b'\0')

        # Combine header and data
        binary = header + data

        # Pad binary with zeros
        binary = binary.ljust(DSF.globals["DCB_MAX_SIZE"], b'\0')

        # Calculate CRC on entire DCB binary
        crc = DSF.bytes_crc32(binary)

        # Replace placeholder CRC
        binary = struct.pack("<I", crc) + binary[4:]
        
        # Print CRC and size information
        print("[INFO] CRC = 0x{:08X}, {:d} bytes free, {:d} of {:d} used".format(crc, DSF.globals["DCB_MAX_SIZE"] - size, size, DSF.globals["DCB_MAX_SIZE"]))

        # Write to file
        #output_filename = "../../ddrsns/target/SDM845_830/settings/release/output_" + sheetname + '_' + re.sub('.xlsx','',(self.filename).split('SDM845_')[-1]) + ".bin"
        if output is None:
            output = DSF.globals["release"]
        with open(os.path.join(output, filename), "wb") as f:
            f.write(binary)

        # Write binaries generated for debug
        if DEBUG:
            for bin in DSF.binaries_ordered_by_header_index:
                # Write to file
                with open("output_{:s}.bin".format(bin), "wb") as f:
                    f.write(self.dcbBinDict[bin])

    def PopulateAbsDict(self, hwio_cmm):
        """
        Populates absDict from ABS sheet in DSF XL file.
        """
        # Populate hwio_dict with register name as key and address as value from HWIO file
        if len(self.absDict) != 0:
            return
        
        p = re.compile(r'^y\.create\.l HWIO_(.+)_ADDR A:0x(.+)$')
        with open(hwio_cmm,"r") as hwio_cmm_fp:
            for hwio_element in hwio_cmm_fp:
                m = p.match(hwio_element)
                if m:
                    self.absDict[m.group(1)] = int(m.group(2), 16)

        # Create scorecard dictionary with count initialized to zero
        scorecardDict = {}
        for block in self.autoDict.keys():
            scorecardDict[block] = 0

        # Create dictionary to map prefix to block
        prefix_to_block_dict = {}
        for block in self.autoDict.keys():
            for prefix in self.autoDict[block]['CMM HWIO Prefixes']:
                prefix_to_block_dict[prefix] = block

        # Check each register for prefix match and increment count for associated block
        prefixes_sorted_from_longest_to_shortest = sorted(prefix_to_block_dict.keys(), key=lambda obj: (len(obj)), reverse=True)
        for reg in self.absDict:
            for prefix in prefixes_sorted_from_longest_to_shortest:
                if prefix in reg:
                    scorecardDict[prefix_to_block_dict[prefix]] += 1
                    break

    def PopulateFreqDict(self):
        freqSheet = self.dsfWorkbook.sheet_by_name("STRUCTs")
        headerArray = freqSheet.row_values(0)

        # Obtain column index for each header
        colIdxDict = {}
        for colIdx in range(len(headerArray)):
            colIdxDict[headerArray[colIdx]] = colIdx

        # Read data from AUTO sheet
        for rowIdx in range(1, freqSheet.nrows):
            rowArray = freqSheet.row_values(rowIdx)
            # Create dictionary for block name
            freq = rowArray[colIdxDict['Frequency']]
            block = rowArray[colIdxDict['Block Name']]
            reg = rowArray[colIdxDict['Register']]
            if freq not in self.freqDict:
                self.freqDict[freq] = {}
            if block not in self.freqDict[freq]:
                self.freqDict[freq][block] = {}

            reg = DSF.ConvertIndexRegisterName(reg)
            self.freqDict[freq][block][reg] = {}

            # Populate dictionary for block name
            self.freqDict[freq][block][reg]["Recommended Settings"] = int(rowArray[colIdxDict['Recommended Setting']], 16)
            self.freqDict[freq][block][reg]["PoR Source"] = int(rowArray[colIdxDict['PoR Source']], 16)

    def PopulateConfigmonDict(self,freq):
        self.PopulateFreqDict()
        self.PopulateRegDictFromSecondarySheets()

        # Merge DCB regDict with per frequency configmon settings
        #for freq in self.freqDict:
        import copy
        self.configmonDict[freq] = copy.deepcopy(self.regDict)
        for block in self.freqDict[freq]:
            for reg in self.freqDict[freq][block]:
                configmon_mask = self.configmonDict[freq][block][reg]["PoR Source"]
                freq_mask = self.freqDict[freq][block][reg]["PoR Source"]
                overlap_mask = configmon_mask & freq_mask
                merged_mask = configmon_mask | freq_mask
                self.configmonDict[freq][block][reg]["PoR Source"] = merged_mask
                if merged_mask != 0xFFFFFFFF:
                    print ("[WARNING] Incomplete mask : {:s} {:X} = {:X} | {:X}".format(reg, merged_mask, configmon_mask, freq_mask))
                if overlap_mask != 0x0:
                    print ("[WARNING] Mask Conflict : {:s} {:X} = {:X} | {:X}".format(reg, overlap_mask, configmon_mask, freq_mask))
                    self.configmonDict[freq][block][reg]["PoR Source"] = 0x0

                masked_configmon_value = self.configmonDict[freq][block][reg]["Recommended Settings"] & self.configmonDict[freq][block][reg]["PoR Source"]
                masked_freq_value = self.freqDict[freq][block][reg]["Recommended Settings"] & self.freqDict[freq][block][reg]["PoR Source"]
                self.configmonDict[freq][block][reg]["Recommended Settings"] = masked_configmon_value | masked_freq_value

    def helperfunction(self, freq):
        for block in self.freqDict[freq]:
            for reg in self.freqDict[freq][block]:
                for cmm in self.autoDict[block]['CMM HWIO Prefixes']:
                    if self.autoDict[block]["Remove Prefix"] == 0:
                        reg_with_cmm = reg.replace(self.autoDict[block]["Summary Prefix"], cmm, 1)
                    else:
                        reg_with_cmm = cmm + "_" + reg
                    self.configmonDict[freq][block][reg]['Configmon Settings'][cmm] = self.t32api.readRegister(self.absDict[reg_with_cmm])#0x0 #read from T32
                    self.configmonDict[freq][block][reg]['Configmon Settings']['Delta'] = 0

    def SuspectCount(self, freq):
        self.helperfunction(freq)
        for block in self.freqDict[freq]:
            for reg in self.freqDict[freq][block]:
                for cmm in self.autoDict[block]['CMM HWIO Prefixes']:
                    #obtain Absolute address
                    if self.autoDict[block]["Remove Prefix"] == 0:
                        reg_with_cmm = reg.replace(self.autoDict[block]["Summary Prefix"], cmm, 1)
                    else:
                        reg_with_cmm = cmm + "_" + reg

                    #determine if suspect instead of regDict use configmonDict
                    recomm_value = self.freqDict[freq][block][reg]["Recommended Settings"]
                    mask = self.freqDict[freq][block][reg]["PoR Source"]
                    value_read = self.configmonDict[freq][block][reg]['Configmon Settings'][cmm]
                    masked_value_read = value_read & mask
                    masked_recomm_value = recomm_value & mask
                    self.configmonDict[freq][block][reg]['Configmon Settings']["Delta"] = (masked_value_read ^ masked_recomm_value)
                    if self.configmonDict[freq][block][reg]['Configmon Settings']["Delta"] != 0:
                        self.scorecardDict[block][freq]["Suspect"] += 1

    def UpdateScorecardSheet(self):
        # Function required: If dict is empty scorecardsheet is empty
        self.out_xl = px.load_workbook(self.filename, keep_vba=True)
        scorecardsheet = self.out_xl.get_sheet_by_name('SCORECARD')
        idx=self.out_xl.get_index(scorecardsheet)
        self.out_xl.remove_sheet(scorecardsheet)
        scorecardsheet = self.out_xl.create_sheet('SCORECARD', idx)
        scorecardsheet.sheet_properties.tabColor= "FF002060"

        headerArray = ["Block Name", "Design Count", "Coverage", "Excluded"]
        freq_sorted = sorted(self.freqDict.keys(), key=lambda obj: obj)
        block_sorted = sorted(self.scorecardDict.keys(), key=lambda obj: len(obj), reverse=True)
        for freq in freq_sorted:
            headerArray.append(freq)
        scorecardsheet.append(headerArray)

        rowArray = []
        for block in block_sorted:
            rowArray = []
            rowArray.append(block)
            rowArray.append(self.scorecardDict[block]["Design Count"])
            rowArray.append(self.scorecardDict[block]["Coverage"])
            rowArray.append(self.scorecardDict[block]["Excluded"])
            for freq in freq_sorted:
                rowArray.append(self.scorecardDict[block][freq]["Suspect"])
            scorecardsheet.append(rowArray)

        headerArray = ["Frequency","Block Name", "Absolute Address", "Register", "Recommended Settings", "PoR Source","Read Value", "Delta"]
        scorecardsheet.append(headerArray)
        for freq in self.freqArray:
            for block in self.freqDict[freq]:
                for reg in self.freqDict[freq][block]:
                    for cmm in self.autoDict[block]["CMM HWIO Prefixes"]:
                        if self.autoDict[block]["Remove Prefix"] == 0:
                            reg_with_cmm = reg.replace(self.autoDict[block]["Summary Prefix"], cmm, 1)
                        else:
                            reg_with_cmm = cmm + "_" + reg

                        if self.configmonDict[freq][block][reg]['Configmon Settings']["Delta"] != 0:
                            row_array = []
                            row_array.append(freq)
                            row_array.append(block)
                            row_array.append(str(hex(self.absDict[reg_with_cmm])))
                            row_array.append(reg_with_cmm)
                            row_array.append(str(hex(self.freqDict[freq][block][reg]["Recommended Settings"])))
                            row_array.append(str(hex(self.freqDict[freq][block][reg]["PoR Source"])))
                            row_array.append(str(hex(self.configmonDict[freq][block][reg]['Configmon Settings'][cmm])))
                            row_array.append(str(hex(self.configmonDict[freq][block][reg]['Configmon Settings']["Delta"])))
                            scorecardsheet.append(row_array)

    def RunConfigmon(self, force = False):
        import openpyxl
        import xlrd
        dsf_xl_sheet = self.dsfWorkbook.sheet_by_name('STRUCT')
        freq_header = []
        freq_header = dsf_xl_sheet.row_values(2)
        ip_addr = input("Enter ip address of host and pressing Enter: ")
        self.t32api = t32api.t32api(str(ip_addr), 20000) #ip required
        connected = self.t32api.connect(str(ip_addr), "20000")
        self.PopulateAbsDict(DSF.globals["target"] + "/hwio/hwio_v1.cmm")
        self.PopulateAutoDict()
        self.PopulateFreqDict()
        self.t32api.first_change_in_freq()
        self.freqArray = []
        freq_len = len(self.freqDict.keys()) + 1
        while True:
            print("Enter any of the below frequencies:")
            print("all frequencies = [0]")
            print(str(freq_header[5]) + " = [1]")
            print(str(freq_header[6]) + " = [2]")
            print(str(freq_header[7]) + " = [3]")
            print(str(freq_header[8]) + " = [4]")
            print(str(freq_header[9]) + " = [5]")
            print(str(freq_header[10]) + " = [6]")
            print(str(freq_header[11]) + " = [7]")
            print(str(freq_header[12]) + " = [8]")
            print(str(freq_header[13]) + " = [9]")
            print(str(freq_header[14]) + " = [10]")
            print(str(freq_header[15]) + " = [11]")
            print("Exit and Store configmon = [12]")
            freq_index = input("Enter any of the options above:")

            if freq_index != "0" and freq_index != "12":
                cmd = self.t32api.CalculateClkFreq(freq_index)
                freq1 = cmd[2]
                freq = freq1[-11:-1]
                freq = int(freq, 16)
                freq = freq/1000
                if freq not in self.freqArray:
                    self.freqArray.append(freq)
                self.PopulateConfigmonDict(freq)
                self.PopulateScorecardDict(freq)
                self.SuspectCount(freq)

            if freq_index == "0":
                for i in range(freq_len):
                    cmd = self.t32api.CalculateClkFreq(i)
                    freq1 = cmd[2]
                    freq = freq1[-11:-1]
                    freq = int(freq, 16)
                    freq = freq/1000
                    if freq not in self.freqArray:
                        self.freqArray.append(freq)
                        self.PopulateConfigmonDict(freq)
                        self.PopulateScorecardDict(freq)
                        self.SuspectCount(freq)
                cmd = self.t32api.break_script()
                break
            if freq_index == "12":
                cmd = self.t32api.CalculateClkFreq(freq_index)
                cmd = self.t32api.break_script()
                break

        self.UpdateScorecardSheet()
        self.out_xl.save(self.filename)
        self.t32api.disconnect()

    def UpdateDCCsheet(self):
        import openpyxl
        self.PopulateAutoDict()
        self.PopulateAbsDict(DSF.globals["target"] + "/hwio/hwio_v1.cmm")
        self.PopulateRegDictFromSecondarySheets()
        prefix_to_block_dict = {}
        self.out_xl = openpyxl.load_workbook(self.filename, keep_vba=True)
        dcc_sheet = self.out_xl.get_sheet_by_name('DCC')
        idx=self.out_xl.get_index(dcc_sheet)
        self.out_xl.remove_sheet(dcc_sheet)
        dcc_sheet = self.out_xl.create_sheet('DCC', idx)
        dcc_sheet.sheet_properties.tabColor= "FFC00000"

        headerArray = ["Block Name", "Absolute Address", "Register", "Clock Switch Hang", "Traffic Stall"]
        dcc_sheet.append(headerArray)
        for block in self.autoDict:
            for prefix in self.autoDict[block]['CMM HWIO Prefixes']:
                prefix_to_block_dict[prefix] = block
        prefixes_sorted_from_longest_to_shortest = sorted(prefix_to_block_dict.keys(), key=lambda obj: (len(obj)), reverse=True)

        for block in self.regDict:
            for reg in self.regDict[block]:
                if self.regDict[block][reg]["Clock Switch Hang"] or self.regDict[block][reg]["Traffic Stall"]:
                    for cmm in self.autoDict[block]["CMM HWIO Prefixes"]:
                        if self.autoDict[block]["Remove Prefix"] == 0:
                            reg_with_cmm = reg.replace(self.autoDict[block]["Summary Prefix"], cmm, 1)
                        else:
                            reg_with_cmm = cmm + "_" + reg
                        row_array = []
                        row_array.append(block)
                        row_array.append(hex(self.absDict[reg_with_cmm]))
                        row_array.append(reg_with_cmm)
                        row_array.append(self.regDict[block][reg]["Clock Switch Hang"])
                        row_array.append(self.regDict[block][reg]["Traffic Stall"])
                        if type(self.regDict[block][reg]["Clock Switch Hang"]) == type('a'):
                            row_array.append("str")
                        if type(self.regDict[block][reg]["Traffic Stall"]) == type('a'):
                            row_array.append("str")

                        dcc_sheet.append(row_array)

        self.out_xl.save(self.filename)

    def PopulateScorecardDict(self, freq):
        self.PopulateAutoDict()

        # Create scorecard dictionary with count initialized to zero
        if len(self.scorecardDict) != 0:
            return

        for block in self.autoDict.keys():
            self.scorecardDict[block] = {}
            for freq in self.freqDict:
                self.scorecardDict[block][freq] = {}
                self.scorecardDict[block]["Design Count"] = 0
                self.scorecardDict[block]["Coverage"] = 0
                self.scorecardDict[block]["Excluded"] = 0
                self.scorecardDict[block][freq]["Suspect"] = 0

        # Create dictionary to map prefix to block
        prefix_to_block_dict = {}
        for block in self.autoDict.keys():
            for prefix in self.autoDict[block]['CMM HWIO Prefixes']:
                prefix_to_block_dict[prefix] = block

        # Check each register for prefix match and increment count for associated block
        prefixes_sorted_from_longest_to_shortest = sorted(prefix_to_block_dict.keys(), key=lambda obj: (len(obj)), reverse=True)
        for reg in self.absDict:
            for prefix in prefixes_sorted_from_longest_to_shortest:
                if prefix in reg:
                    self.scorecardDict[prefix_to_block_dict[prefix]]["Design Count"] += 1
                    break
        for block in self.regDict:
            for reg in self.regDict[block]:
                self.scorecardDict[block]["Coverage"] += len(self.autoDict[block]["CMM HWIO Prefixes"])
                if self.regDict[block][reg]["PoR Source"] != 0xFFFFFFFF:
                    self.scorecardDict[block]["Excluded"] += len(self.autoDict[block]["CMM HWIO Prefixes"])

    def UpdateRelativeAddresses(self):
        import openpyxl
        self.PopulateAutoDict()
        #Opening DSF XL file in write mode

        self.out_xl = openpyxl.load_workbook(self.filename, keep_vba=True)
        # Open REL sheet in DSF XL file

        rel_sheet = self.out_xl.get_sheet_by_name('REL')
        #print(rel_sheet.sheet_properties.tabColor)
        idx=self.out_xl.get_index(rel_sheet )
        self.out_xl.remove_sheet(rel_sheet)
        rel_sheet = self.out_xl.create_sheet('REL',idx)
        #rel_sheet.sheet_properties.tabColor= "FFFFC000"
        #rel_sheet = self.out_xl.get_sheet_by_name('REL')
        rel_sheet.append(['Relative address', 'Register'])

        # Extract relative addresses from each *_hwioreg.h file
        visitheader=[]
        for block in self.autoDict:
            if self.autoDict[block]['Header'] not in visitheader:
                reg_hwfilename = DSF.globals["COMMON"] + self.autoDict[block]['Header']
                visitheader.append(self.autoDict[block]['Header'])
                with open(reg_hwfilename,'r') as regfile:
                    pre_m=0
                    pre_n=0
                    prev_ltxt=[]
                    rev_mn=0
                    for line in regfile:
                        addr=''
                        reg=''
                        line=(re.sub('[\(\)\{\}<>]', '  ',line)).strip("#\n\t")#.split("., ")
                        ltxt=line.split(' ')
                        ltxt=list(filter(None, ltxt))
                        if ltxt!=[] and ltxt[0] == "define":
                            if ltxt[1].find("_ADDR")!=-1 and ltxt[1][-5:]=="_ADDR":
                                if len(ltxt)==4:
                                    reg= re.sub('HWIO_','',ltxt[1])
                                    reg= reg[:-5]
                                    addr= ltxt[3].split('+')[-1][2:]
                                    if len(addr)<8:
                                        addr=(8-len(addr))*'0'+addr
                                    if int(self.autoDict[block]["Remove Prefix"]):
                                        reg = re.sub(self.autoDict[block]["C Code HWIO Prefix"]+'_', '', reg)
                                    rel_sheet.append([addr, reg])
                                elif len(ltxt)==5:
                                    pre_n=1
                                    pre_m=0
                                    prev_ltxt = ltxt
                                elif len(ltxt)==6:
                                    if ltxt[3] == "n,":
                                        tmp_line= re.sub('m','t',line)
                                        tmp_line= re.sub('n','m',tmp_line)
                                        tmp_line= re.sub('t','n',tmp_line)
                                        ltxt1=tmp_line.split(' ')
                                        ltxt=list(filter(None, ltxt1))
                                        rev_mn=1
                                    pre_n=1
                                    pre_m=1
                                    prev_ltxt = ltxt

                            elif pre_n==1 and pre_m==0 and (ltxt[1].find("_MAXm")!=-1 or ltxt[1].find("_MAXn")!=-1):
                                reg1= re.sub('HWIO_','',prev_ltxt[1])
                                reg1= reg1[:-5]
                                for i in range(int(ltxt[-1])+1):

                                    if reg1.find('n')!=-1:
                                        reg= re.sub('n',str(i),reg1)
                                    else:
                                        reg= re.sub('m',str(i),reg1)
                                    tmp_addr_list= prev_ltxt[-1].split('+')
                                    if len(tmp_addr_list) == 2:
                                        addr= hex((int((prev_ltxt[-1].split('+'))[-1].split('*')[0],16)*i))[2:]
                                    else :
                                        addr= hex(int(prev_ltxt[-1].split('+')[1],16)+(int((prev_ltxt[-1].split('+'))[-1].split('*')[0],16)*i))[2:]
                                    if len(addr)<8:
                                        addr=(8-len(addr))*'0'+addr


                                    if int(self.autoDict[block]["Remove Prefix"]):
                                        reg = re.sub(self.autoDict[block]["C Code HWIO Prefix"]+'_', '', reg)
                                    rel_sheet.append([addr, reg])
                                prev_n=0
                                prev_ltxt=''
                            elif pre_n==1 and pre_m==1 and ltxt[1].find("_MAXm")!=-1:
                                mval=int(ltxt[-1])
                            elif pre_n==1 and pre_m==1 and ltxt[1].find("_MAXn")!=-1:
                                reg1= re.sub('HWIO_','',prev_ltxt[1])
                                reg1= reg1[:-5]
                                if rev_mn==1:
                                    mval=int(ltxt[-1])
                                    nextline = regfile.__next__()
                                    nextline=(re.sub('[\(\)\{\}<>]', '  ',nextline)).strip("#\n\t")
                                    nextltxt=list(filter(None, nextline.split(' ')))
                                    ltxt=nextltxt
                                for i in range(mval+1):
                                    for j in range(int(ltxt[-1])+1):
                                        reg= re.sub('m',str(i),reg1)
                                        reg= re.sub('n',str(j),reg)
                                        tmp_addr_list= prev_ltxt[-1].split('+')
                                        if len(tmp_addr_list) == 3:
                                            addr= hex((int((prev_ltxt[-1].split('+'))[1].split('*')[0],16)*i)+(int((prev_ltxt[-1].split('+'))[-1].split('*')[0],16)*j))[2:]
                                        else :
                                            addr= hex(int(prev_ltxt[-1].split('+')[1],16)+(int((prev_ltxt[-1].split('+'))[2].split('*')[0],16)*i)+(int((prev_ltxt[-1].split('+'))[-1].split('*')[0],16)*j))[2:]
                                        if len(addr)<8:
                                            addr=(8-len(addr))*'0'+addr

                                        if int(self.autoDict[block]["Remove Prefix"]):
                                            reg = re.sub(self.autoDict[block]["C Code HWIO Prefix"]+'_', '', reg)
                                        rel_sheet.append([addr, reg])
                                pre_n=0
                                pre_m=0
                                prev_ltxt=0
                                rev_mn=0


        self.out_xl.save(self.filename)
        self.RunSwcMacros()

    def PopulateRelDict(self, force=False):
        """
        Populates relDict from REL sheet in DSF XL file.
        """
        if len(self.relDict) != 0 and force == False:
            return

        #Open REL sheet in DSF XL file
        reladdr_sheet=self.dsfWorkbook.sheet_by_name('REL')

        row_count= reladdr_sheet.nrows

        # Populate Relative Address Dict
        for i in range(1,row_count):
            text = reladdr_sheet.row_values(i)
            self.relDict[text[1]]=text[0]

    def RunSwcMacros(self):
        import win32com.client
        xl = None
        xl = win32com.client.Dispatch("Excel.Application")
        xl.Workbooks.Open(Filename=os.path.abspath(self.filename))
        xl.Application.Run("HeaderMacro")
        xl.Workbooks(1).Close(SaveChanges=1)
        xl.Application.Quit()
        del xl

    def GenerateExternalDsfXl(self):
        import win32com.client
        xl = None
        xl = win32com.client.Dispatch("Excel.Application")
        xl.Workbooks.Open(Filename=os.path.abspath(self.filename))
        try:
            xl.Application.Run("ProtectWorkbook")
        except:
            print("[WARNING] ProtectWorkbook macro not found")
        xl.Workbooks(1).Close(SaveChanges=1)
        xl.Application.Quit()
        del xl

    def PopulateDcbBinDictFromBinarySheets(self):
        # Open sheet with DCB binary data for structures
        sheet = self.dsfWorkbook.sheet_by_name('STRUCTb')

        # Read header from sheet
        headerArray = sheet.row_values(0)

        # Obtain column index for each header
        colIdxDict = {}
        for colIdx in range(len(headerArray)):
            colIdxDict[headerArray[colIdx]] = colIdx

        # Temporary items
        valueDict = {} # three dimensional array <block><band><value>
        active = ''
        count = 0

        # Read values from sheet
        for rowIdx in range(1, sheet.nrows):
            rowArray = sheet.row_values(rowIdx)

            # Check Block Name
            block = rowArray[colIdxDict['Block Name']]
            if block not in valueDict:
                valueDict[block] = []
                count = 0
                # Delete to replace content in DCB
                if block in self.dcbBinDict:
                    del self.dcbBinDict[block]

            # Check Structure/Array
            structure = rowArray[colIdxDict['Structure/Array']]
            if structure != active:
                valueDict[block].append([[] for i in range(len(rowArray) - colIdxDict[0] + 1)])
                active = structure
                count += 1

            # Place values into bands
            band = 0
            for colIdx in range(colIdxDict[0], len(rowArray)):
                valueDict[block][count-1][band].append(rowArray[colIdx])
                band += 1

        # Turn values into bytes
        for block in valueDict:
            if block not in self.dcbBinDict:
                self.dcbBinDict[block] = bytes()

            for structure in valueDict[block]:
                for band in structure:
                    for value in band:
                        if len(value) == 2:
                            self.dcbBinDict[block] += struct.pack('B', int(value, 16))
                        elif len(value) == 4:
                            self.dcbBinDict[block] += struct.pack('H', int(value, 16))
                        elif len(value) == 8:
                            self.dcbBinDict[block] += struct.pack('I', int(value, 16))

    def PopulateTrainingRegister(self, force=False):
        if len(self.trainingRegDict) != 0 and force == False:
            return
        block_name = ''
        count=0
        with open(DSF.globals["target"] + '/src/ddr_phy_restore.c') as regfile:
            for line in regfile:
                line=(re.sub('[\(\)\{\}<>]', '  ',line)).strip("#\n\t")
                ltxt=line.split(' ')
                ltxt=list(filter(None, ltxt))
                if ltxt!=[] and ltxt[0] == "uint32" and ltxt[1].find('[]'):
                    block_name =  (re.sub('\[\]', '',ltxt[1]))
                    self.trainingRegDict[block_name]={}
                    count = 0
                elif ltxt!=[] and block_name != '':
                    if ltxt[0]=='0x0':
                        block_name = ''
                    elif len(ltxt)>=3:
                        self.trainingRegDict[block_name][ltxt[2]]={}
                        self.trainingRegDict[block_name][ltxt[2]]['Index in Stack'] = count
                        count= count + 1

    def UpdateTrainingsheet(self):
        import openpyxl
        self.PopulateTrainingRegister()
        self.out_xl = openpyxl.load_workbook(self.filename,keep_vba=True)
        phy_sheet = self.out_xl.get_sheet_by_name('TRAINING')
        idx=self.out_xl.get_index(phy_sheet)
        self.out_xl.remove_sheet(phy_sheet)
        phy_sheet = self.out_xl.create_sheet('TRAINING', idx)
        phy_sheet.sheet_properties.tabColor='FFFFC000'

        with open(DSF.globals["COMMON"] + 'ddr_phy_seq_hwioreg.h') as regfile:
            for line in regfile:
                line=(re.sub('[\(\)\{\}<>]', '  ',line)).strip("#\n\t")#.split("., ")
                ltxt=line.split(' ')
                ltxt=list(filter(None, ltxt))
                if ltxt!=[] and ltxt[0] == "define":
                    nextline = regfile.__next__()
                    nextline=(re.sub('[\(\)\{\}<>]', '  ',nextline)).strip("#\n\t")
                    nextltxt=list(filter(None, nextline.split(' ')))
                    for block in self.trainingRegDict:
                        for regkey in self.trainingRegDict[block]:
                            if ltxt[1].find(regkey) !=-1 and (ltxt[1].split(regkey))[1] == "_ADDR":
                                self.trainingRegDict[block][regkey]['ADDR'] = ltxt[3].split('+')[1]
                            elif ltxt[1].find(regkey) !=-1 and ltxt[1].find("_BMSK") !=-1:
                                flagtxt =(ltxt[1].split(regkey))[1].split('_BMSK')
                                self.trainingRegDict[block][regkey][flagtxt[0][1:]] ={}
                                self.trainingRegDict[block][regkey][flagtxt[0][1:]]['BMSK'] = ltxt[2]
                                if nextltxt[1].find(regkey) !=-1 and nextltxt[1].find("_SHFT") !=-1:
                                    self.trainingRegDict[block][regkey][flagtxt[0][1:]]['SHFT'] = nextltxt[2]


        phy_sheet.append(["Block","Index in Stack","Address","self.trainingRegDictister","Field","Bit Field Mask","Shift"])
        for block in self.trainingRegDict:
            #row =[block]
            #phy_sheet.append(row)
            for regkey in self.trainingRegDict[block]:
            #    row = ["",self.trainingRegDict[block][regkey]['Index in Stack'],self.trainingRegDict[block][regkey]["ADDR"],regkey,""]
            #    phy_sheet.append(row)
                for field in self.trainingRegDict[block][regkey]:
                    row =[block,self.trainingRegDict[block][regkey]['Index in Stack'],self.trainingRegDict[block][regkey]["ADDR"],regkey,field]
                    if field != "ADDR" and field != 'Index in Stack':

                        row.append(self.trainingRegDict[block][regkey][field]['BMSK'])
                        row.append(self.trainingRegDict[block][regkey][field]['SHFT'])
                        phy_sheet.append(row)


        self.out_xl.save(self.filename)

    def PopulateTrainingDictionary(self, force=False):
        if len(self.trainingSheetDict) != 0 and force == False:
            return

        training_sheet = self.dsfWorkbook.sheet_by_name('TRAINING')
        row_count = training_sheet.nrows
        self.trainingSheetDict={}
        prev_block = ''
        prev_reg = ''
        for i in range(row_count):
            row_text = training_sheet.row_values(i)
            if row_text[0] != '':
                self.trainingSheetDict[row_text[0]] = {}
                prev_block = row_text[0]
            elif row_text[1] != '':
                self.trainingSheetDict[prev_block][row_text[3]] = {}
                prev_reg = row_text[3]
                self.trainingSheetDict[prev_block][row_text[3]]['Index in Stack'] = int(row_text[1])
                self.trainingSheetDict[prev_block][row_text[3]]["ADDR"] = row_text[2]
            else :
                self.trainingSheetDict[prev_block][prev_reg][row_text[4]]={}
                self.trainingSheetDict[prev_block][prev_reg][row_text[4]]['BMSK'] = row_text[5]
                self.trainingSheetDict[prev_block][prev_reg][row_text[4]]['SHFT'] = row_text[6]

    def GenerateTrainingValues(self, binary_file):
        import numpy
        self.PopulateTrainingDictionary()

        #generate null terminated arrays in dcb
        training_sheet = self.dsfWorkbook.sheet_by_name('TRAINING')

        #reading binary values from bin file
        with open(binary_file, mode='rb') as file:
            fileContent = file.read()

        file_size = len(fileContent)
        stack_array =[]

        for i in range(int(file_size/4)):
            stack_array.append(int(((fileContent[i*4+3])<<24)+((fileContent[i*4+2])<<16)+((fileContent[i*4+1])<<8)+(fileContent[4*i]<<0)))

        Array = {}
        offset = int((int('0x52AC',16)+int('0x14',16)+int('0x6C',16))/4)

        target1 =open('output_stack.txt','w')
        target1.truncate()
        for i in range(offset,int(file_size/4)):
            target1.write("%s\t"%hex(fileContent[i*4+3]))
            target1.write("%s\t"%hex(fileContent[i*4+2]))
            target1.write("%s\t"%hex(fileContent[i*4+1]))
            target1.write("%s"%hex(fileContent[i*4+0]))
            target1.write("\t")
            target1.write("%s\n"%hex(stack_array[i]))
        target1.close()
        for blockkey in self.trainingSheetDict:
            Array[blockkey]={}
            for regkey in self.trainingSheetDict[blockkey]:
                Array[blockkey][regkey]={}
                for fieldkey in self.trainingSheetDict[blockkey][regkey]:
                    if fieldkey != 'ADDR' and fieldkey != 'Index in Stack':
                        if blockkey.find('CA_PHY_REGS') !=-1:
                            Array[blockkey][regkey][fieldkey] = numpy.zeros((DSF.globals["NUM_CH"],DSF.globals["NUM_CA_PCH"]))
                            for ch in range(DSF.globals["NUM_CH"]):
                                for ca in range(DSF.globals["NUM_CA_PCH"]):

                                    Array[blockkey][regkey][fieldkey][ch][ca] = (stack_array[int(offset+ch*DSF.globals["NUM_CA_PCH"]*DSF.globals["NUM_CA_TRAINING_STACK"]+ca*DSF.globals["NUM_CA_TRAINING_STACK"]+self.trainingSheetDict[blockkey][regkey]['Index in Stack'])] & int(self.trainingSheetDict[blockkey][regkey][fieldkey]['BMSK'],16))>>int(self.trainingSheetDict[blockkey][regkey][fieldkey]['SHFT'],16)

                        elif blockkey.find('DQ_PHY_REGS') !=-1:

                            dq_offset = offset + DSF.globals["NUM_CA_PCH"]*DSF.globals["NUM_CH"]*DSF.globals["NUM_CA_TRAINING_STACK"]
                            Array[blockkey][regkey][fieldkey] = numpy.zeros((DSF.globals["NUM_CH"],DSF.globals["NUM_DQ_PCH"]))
                            for ch in range(DSF.globals["NUM_CH"]):
                                for dq in range(DSF.globals["NUM_DQ_PCH"]):

                                    Array[blockkey][regkey][fieldkey][ch][dq] = (stack_array[int(dq_offset+ch*DSF.globals["NUM_DQ_PCH"]*DSF.globals["NUM_DQ_TRAINING_STACK"]+dq*DSF.globals["NUM_DQ_TRAINING_STACK"]+self.trainingSheetDict[blockkey][regkey]['Index in Stack'])] & int(self.trainingSheetDict[blockkey][regkey][fieldkey]['BMSK'],16))>>int(self.trainingSheetDict[blockkey][regkey][fieldkey]['SHFT'],16)


        target =open('output.txt','w')
        target.truncate()
        for blockkey in self.trainingSheetDict:
            target.write("\n")
            target.write(blockkey)
            for regkey in self.trainingSheetDict[blockkey]:
                target.write("\n")
                target.write(regkey)
                for fieldkey in self.trainingSheetDict[blockkey][regkey]:
                    target.write("\n")
                    target.write(fieldkey)
                    if fieldkey != 'ADDR' and fieldkey!= 'Index in Stack' and blockkey.find('CA_PHY_REGS') !=-1:
                        target.write("\nBMSK     %s\n" %self.trainingSheetDict[blockkey][regkey][fieldkey]["BMSK"])
                        target.write("SHFT     %s\n" %self.trainingSheetDict[blockkey][regkey][fieldkey]["SHFT"])
                        for ch in range(DSF.globals["NUM_CH"]):
                            for ca in range(DSF.globals["NUM_CA_PCH"]):
                                target.write("%u    " % Array[blockkey][regkey][fieldkey][ch][ca])
                            target.write("\n")
                    elif fieldkey != 'ADDR' and fieldkey!= 'Index in Stack' and blockkey.find('DQ_PHY_REGS') !=-1:
                        target.write("\nBMSK     %s\n" %self.trainingSheetDict[blockkey][regkey][fieldkey]["BMSK"])
                        target.write("SHFT     %s\n" %self.trainingSheetDict[blockkey][regkey][fieldkey]["SHFT"])
                        for ch in range(DSF.globals["NUM_CH"]):
                            for dq in range(DSF.globals["NUM_DQ_PCH"]):
                                target.write("%u    " % Array[blockkey][regkey][fieldkey][ch][dq])
                            target.write("\n")
                    else :
                        target.write(" %s    " % self.trainingSheetDict[blockkey][regkey][fieldkey])
                        target.write("\n")
                target.write("\n")
        target.close()
        return Array

    def PopulateDcbBinDictPhyRegs(self):
        self.PopulateTrainingDictionary()


        for phy in self.trainingSheetDict:
            self.dcbBinDict[phy] = bytes()
            registers_sorted_by_index = sorted(self.trainingSheetDict[phy].keys(), key=lambda obj: self.trainingSheetDict[phy][obj]['Index in Stack'])
            for reg in registers_sorted_by_index:
                self.dcbBinDict[phy] += struct.pack("<I", int(self.trainingSheetDict[phy][reg]['ADDR'], 16))
            if phy == "CA_PHY_REGS":
                self.dcbBinDict[phy] = self.dcbBinDict[phy].ljust(DSF.globals["NUM_CA_TRAINING_STACK"], b'\0')
            else:
                self.dcbBinDict[phy] = self.dcbBinDict[phy].ljust(DSF.globals["NUM_DQ_TRAINING_STACK"], b'\0')

    def PopulateBaseDict(self, force=False):
        if len(self.baseDict) != 0 and force == False:
            return

        visitheader=[]
        for block in self.autoDict:
            if self.autoDict[block]['Header'] not in visitheader:
                reg_base = self.autoDict[block]['Header'].replace("reg","base")
                reg_hwfilename = DSF.globals["COMMON"] + reg_base
                visitheader.append(self.autoDict[block]['Header'])
                with open(reg_hwfilename,'r') as base_fp:
                    base_element = base_fp.readline()
                    for base_element in base_fp:
                        base_element = (re.sub('[\(\)\{\}<>]', '  ',base_element)).strip("#\n    ")#.split("., ")
                        base_str = base_element.split(' ')
                        base_str = list(filter(None, base_str))
                        if base_str != [] and base_str[0] == "define":
                            if "_SEQ_BASE_H__" not in base_str[1]:
                                base_hex = hex(int(base_str[2], 16))
                                self.baseDict[base_str[1]] = base_hex

    def PopulateRegDictfromDSFSheet(self, sheet, start):
        Output_Dict={}
        colIdxDict = {}
        if sheet.nrows ==0:
            return Output_Dict
        headerArray = sheet.row_values(start)
        for colIdx in range(len(headerArray)) :
            if headerArray[colIdx] == 'Register':
                regIdx = colIdx
        # Read data from AUTO sheet
        col_count = sheet.ncols
        for rowIdx in range(start+1, sheet.nrows):
            rowArray = sheet.row_values(rowIdx)
            blockName = rowArray[0]
            if blockName not in Output_Dict:
                Output_Dict[blockName] = {}
            if '<' in rowArray[regIdx]:
                rowArray[regIdx] = DSF.ConvertIndexRegisterName(rowArray[regIdx])
            Output_Dict[blockName][rowArray[regIdx]]={}
            for colIdx in range(1,col_count):
                # Populate dictionary for block name
                if colIdx != regIdx:
                    Output_Dict[blockName][rowArray[regIdx]][headerArray[colIdx]] = rowArray[colIdx]
                elif colIdx == regIdx:
                    Output_Dict[blockName][rowArray[regIdx]]['Index'] = rowIdx
        return Output_Dict

    def PopulateRegDictFromPrimarySheets(self):
        if len(self.regDict) != 0:
            return

        self.PopulateAutoDict()
        self.PopulateBaseDict()
        if self.regDict == {}:
            dsf_xl_sheet = self.dsfWorkbook.sheet_by_name('SWCs')
            self.regDict = self.PopulateRegDictfromDSFSheet(dsf_xl_sheet,len(self.xlfileDict)+1)
        for block in self.regDict:
            for reg in self.regDict[block]:
                for element in self.regDict[block][reg]:
                    if (element in ['Relative Address','Recommended Settings','Default','PoR Source','Clock Switch Hang','Traffic Stall']) and type(self.regDict[block][reg][element]) == type('a'):
                        try :
                            self.regDict[block][reg][element] = int(self.regDict[block][reg][element],16)
                        except:
                            print(block + ' ' + reg + ' ' + element)
                self.regDict[block][reg]['DCB Binary Sub-Block'] = self.autoDict[block]['DCB Binary Sub-Block']
                self.regDict[block][reg]['Configmon Settings'] = {}

                # C Code Register Offset getting added to original relative address
                relative_address = self.regDict[block][reg]['Relative Address']

                if str(self.autoDict[block]['C Code Register Offset']) != "0":
                    relative_address += int(self.baseDict[self.autoDict[block]['C Code Register Offset']], 16)
                else:
                    relative_address += int(self.autoDict[block]['C Code Register Offset'], 16)

                self.regDict[block][reg]['Relative Address'] = relative_address
                
                self.regDict[block][reg]['Merged'] = False

    def PopulateRegDictFromSecondarySheets(self):
        if len(self.regDict) != 0:
            return

        self.PopulateRegDictFromPrimarySheets()

        for sheet in ["ALCs","TIMINGs"]:
            try:
                sheet_to_merge_fp = self.dsfWorkbook.sheet_by_name(sheet)
            except:
                continue
            # Read header from AUTO sheet
            headerArray = sheet_to_merge_fp.row_values(0)

            # Obtain column index for each header
            colIdxDict = {}
            for colIdx in range(len(headerArray)):
                colIdxDict[headerArray[colIdx]] = colIdx

            # Read data from AUTO sheet
            for rowIdx in range(1, sheet_to_merge_fp.nrows):
                rowArray = sheet_to_merge_fp.row_values(rowIdx)

                # Create dictionary for block name
                block = rowArray[colIdxDict['Block Name']]
                reg = rowArray[colIdxDict['Register']]

                reg = DSF.ConvertIndexRegisterName(reg)

                primary_mask_stored = self.regDict[block][reg]['PoR Source']
                primary_value_stored = self.regDict[block][reg]['Recommended Settings']

                secondary_mask_stored = int(rowArray[colIdxDict['PoR Source']], 16)
                try:
                    secondary_value_stored = int(rowArray[colIdxDict['Recommended Settings']], 16)
                except:
                    print(rowArray)
                self.regDict[block][reg]['Recommended Settings'] = (primary_mask_stored & primary_value_stored) | (secondary_mask_stored & secondary_value_stored)
                overlap_mask = primary_mask_stored & secondary_mask_stored
                merged_mask = primary_mask_stored | secondary_mask_stored

                self.regDict[block][reg]['PoR Source'] = merged_mask

                if (merged_mask) != 0xFFFFFFFF:
                    print ("[WARNING] Incomplete mask : {:s} {:X} = {:X} | {:X}".format(reg, merged_mask, primary_mask_stored, secondary_mask_stored))
                if overlap_mask != 0:
                    print ("[WARNING] Mask Conflict : {:s} {:X} = {:X} | {:X}".format(reg, overlap_mask, primary_mask_stored, secondary_mask_stored))
                    self.regDict[block][reg]['PoR Source'] = 0
                    
                # Indicate register setting is the result of a merge
                self.regDict[block][reg]['Merged'] = True


    def PopulateDcbBinDictFromSheets(self):
        self.PopulateDcbBinDictFromBinarySheets()
        self.PopulateRegDictFromSecondarySheets()
        self.PopulateDcbBinDictPhyRegs()

        dcb_blocks = []
        reg_blocks_sorted = sorted(self.regDict.keys(), key=lambda obj: obj)
        for reg_block in reg_blocks_sorted:
            reg_names_sorted = sorted(self.regDict[reg_block].keys(), key=lambda obj: self.regDict[reg_block][obj]['Relative Address'])

            for regName in reg_names_sorted:
                dcb_block = self.regDict[reg_block][regName]['DCB Binary Sub-Block']
                if dcb_block not in self.dcbBinDict:
                    self.dcbBinDict[dcb_block] = bytes()
                    dcb_blocks.append(dcb_block)
                if self.regDict[reg_block][regName]['Recommended Settings'] != self.regDict[reg_block][regName]['Default'] or self.regDict[reg_block][regName]['Merged']:
                    self.dcbBinDict[dcb_block] += struct.pack("<I", self.regDict[reg_block][regName]['Relative Address'])
                    self.dcbBinDict[dcb_block] += struct.pack("<I", self.regDict[reg_block][regName]['Recommended Settings'])

        for dcb_block in dcb_blocks:
            self.dcbBinDict[dcb_block] += struct.pack("<Q", 0)

    def PopulateDictforSWCsheet(self, sheet, input_Dict):
        Output_Dict = input_Dict
        colIdxDict = {}
        headerArray = sheet.row_values(0)
        for colIdx in range(len(headerArray)) :
            if headerArray[colIdx] == 'Register':
                regIdx = colIdx
        # Read data from AUTO sheet
        col_count = sheet.ncols
        blockName = ''
        for rowIdx in range(1, sheet.nrows):
            rowArray = sheet.row_values(rowIdx)
            if rowArray[regIdx]=='' and rowArray[0] != '':
                blockName = rowArray[0]
            if rowArray[regIdx]!='':
                if blockName not in Output_Dict:
                    Output_Dict[blockName] = {}
                register = DSF.ConvertIndexRegisterName(rowArray[regIdx])
                Output_Dict[blockName][register]={}
                for colIdx in range(1,col_count):
                    if colIdx != regIdx:
                    # PHY team requires they use 'Default value' instead of 'Default'
                        if headerArray[colIdx] == 'Default value':
                            Output_Dict[blockName][register]['Default'] = rowArray[colIdx]
                        else:
                            Output_Dict[blockName][register][headerArray[colIdx]] = rowArray[colIdx]
        return Output_Dict

    def UpdateSWCblockDict(self,swc_xl_Dict,version,block):

        if self.regDict[block] == {} or version == 0:
            for reg in swc_xl_Dict[block]:
                if reg not in self.regDict[block]:
                    self.regDict[block][reg] = {}
                for element in swc_xl_Dict[block][reg]:
                    if element in DSF.dsf_xl_columns:
                        self.regDict[block][reg][element]=swc_xl_Dict[block][reg][element]

                if reg in self.relDict:
                    self.regDict[block][reg]['Relative Address'] = self.relDict[reg]
                else:
                    cprefix= self.autoDict[block]['C Code HWIO Prefix'] +'_'
                    register = reg[len(cprefix):]
                    if register in self.relDict:
                        self.regDict[block][reg]['Relative Address'] = self.relDict[register]
                    else:
                        print("[ERROR] %s is not present in REL sheet" %register)
        else :
            for reg in swc_xl_Dict[block]:
                if reg not in self.regDict[block]:
                    self.regDict[block][reg] = {}
                    for element in swc_xl_Dict[block][reg]:
                        if (element in DSF.dsf_xl_columns):
                            self.regDict[block][reg][element]=swc_xl_Dict[block][reg][element]
                else:
                    for element in self.regDict[block][reg]:
                        if (element in DSF.dsf_xl_columns) and (element in swc_xl_Dict[block][reg]):
                            self.regDict[block][reg][element]=swc_xl_Dict[block][reg][element]
                        elif element != 'Relative Address' and element !='Index' and element != 'Default value':
                            print("[ERROR] %s column is not present in %s"%(element,self.autoDict[block]['XL file']))

                if reg in self.relDict:
                    self.regDict[block][reg]['Relative Address'] = self.relDict[reg]
                else:
                    cprefix= self.autoDict[block]['C Code HWIO Prefix'] +'_'
                    register = reg[len(cprefix):]
                    if register in self.relDict:
                        self.regDict[block][reg]['Relative Address'] = self.relDict[register]
                    else:
                        print("[ERROR] %s is not present in REL sheet" %register)

    def SwcDictPrint(self):
        import openpyxl
        self.out_xl = openpyxl.load_workbook(self.filename,keep_vba=True)
        dsf_xl_sheet_output = self.out_xl.get_sheet_by_name('SWCs')
        idx=self.out_xl.get_index(dsf_xl_sheet_output)
        self.out_xl.remove_sheet(dsf_xl_sheet_output)
        dsf_xl_sheet_output = self.out_xl.create_sheet('SWCs',idx)
        dsf_xl_sheet_output.sheet_properties.tabColor='FF92D050'
        for xlfile in self.xlfileDict:
            dsf_xl_sheet_output.append([xlfile,self.xlfileDict[xlfile]])
        dsf_xl_sheet_output.append([])
        dsf_xl_sheet_output.append(DSF.dsf_xl_columns)
        sorted_blockDict = sorted(self.regDict.keys(), key=lambda obj: self.autoDict[obj]['Scorecard Index'])
        for block in sorted_blockDict:
            sorted_regDict = sorted(self.regDict[block].keys(), key=lambda obj: int(self.regDict[block][obj]['Relative Address'],16))
            for reg in sorted_regDict:
                list =['']*len(DSF.dsf_xl_columns)
                for element in self.regDict[block][reg]:
                    if element in DSF.dsf_xl_columns:
                        list[DSF.dsf_xl_columns.index(element)]= self.regDict[block][reg][element]
                list[0] = block
                list[2] = reg
                dsf_xl_sheet_output.append(list)
        self.out_xl.save(self.filename)
        self.RunSwcMacros()

    def UpdateSwcSheet(self):
        self.PopulateAutoDict()
        self.UpdateRelativeAddresses()
        self.PopulateRelDict()

        #reading the SWC sheet in DSF xl file, generating Dict and saving it in Self.regDict
        dsf_xl_sheet = self.dsfWorkbook.sheet_by_name('SWCs')
        self.regDict = {}

        if dsf_xl_sheet.nrows !=0:
            self.regDict = self.PopulateRegDictfromDSFSheet(dsf_xl_sheet,len(self.xlfileDict)+1)
            for i in range(len(self.xlfileDict)):
                xl_file_name = dsf_xl_sheet.row_values(i)[0]
                if xl_file_name in self.xlfileDict:
                    self.xlfileDict[xl_file_name] = float(dsf_xl_sheet.row_values(i)[1])
        update_flag = 0
        bypass = input("Bypasss checking version number of SWC against DSF XL file (Y/N):")
        for xlfile in self.xlfileDict:
            #reading the SWC excel file and getting version
            swc_xl_sheet_filename=DSF.globals["target"] + "/swc/" + xlfile
            try:
                swc_xl = xlrd.open_workbook(swc_xl_sheet_filename)
            except:
                print("Unable to open excel spreadsheet {:s}".format(swc_xl_sheet_filename))
                continue
            swc_xl_sheet = swc_xl.sheet_by_name('Rev.History')
            xl_row_count = swc_xl_sheet.nrows
            swc_xl_version = 0
            for i in range(xl_row_count):
                if swc_xl_sheet.row_values(xl_row_count-i-1)[0] != "":
                    swc_xl_version = float(swc_xl_sheet.row_values(xl_row_count-i-1)[0])
                    break

            #Updating the block if the excel version has changed
            visit_block = []
            if (swc_xl_version > self.xlfileDict[xlfile] and (bypass == 'n' or bypass == 'N')) or bypass == 'y' or bypass == 'Y':
                for block in self.autoDict:
                    if self.autoDict[block]['XL file'] == xlfile:
                        print("[INFO] %s block is updated from the excel file %s"%(block,xlfile))
                        if block not in visit_block:
                            swc_xl_Dict ={}
                            for sheet in self.autoDict[block]['Summary Sheet']:
                                swc_xl_Dict = self.PopulateDictforSWCsheet(swc_xl.sheet_by_name(sheet),swc_xl_Dict)
                            if block not in self.regDict:
                                self.regDict[block] ={}
                            if block in swc_xl_Dict:
                                self.UpdateSWCblockDict(swc_xl_Dict,self.xlfileDict[xlfile],block) #dsf_xl_sheet_Dict =
                            self.xlfileDict[xlfile] = swc_xl_version
                            visit_block.append(block)
                update_flag = 1
        if update_flag == 1:
            self.SwcDictPrint()
        self.RunSwcMacros()

    def Comparesheets(self,filename1,filename2):
        import openpyxl
        self.PopulateAutoDict()
        #reading files
        xlfile_v1 = xlrd.open_workbook(filename1) # 'target/SDM845_830/SDM845_V1_ASIC.xlsx'
        xlfile_v2 = xlrd.open_workbook(filename2) # 'target/SDM845_830/SDM845_V2_RUMI.xlsx'
        output = openpyxl.load_workbook(filename2, keep_vba=True)
        sheets=['SWCs']
        sheet = 'Changes'
        if sheet in output.sheetnames:
            output_changes=output.get_sheet_by_name(sheet)
            output.remove_sheet(output_changes)
        output_changes = output.create_sheet(sheet)


        for sheet in sheets:
            sheet_v1 = xlfile_v1.sheet_by_name(sheet)
            sheet_v1_Dict = self.PopulateRegDictfromDSFSheet(sheet_v1,len(self.xlfileDict)+1)
            sheet_v2 = xlfile_v2.sheet_by_name(sheet)
            sheet_v2_Dict = self.PopulateRegDictfromDSFSheet(sheet_v2,len(self.xlfileDict)+1)
            list = sheet_v2.row_values(1)
            list.insert(0,'Action')
            list.insert(1,'Sheet')
            output_changes.append(list)
            for blockName in sheet_v2_Dict:
                for regName in sheet_v2_Dict[blockName]:
                    flag=0
                    if (blockName not in sheet_v1_Dict) or (regName not in sheet_v1_Dict[blockName]):
                        flag=1
                    else:
                        for element in sheet_v2_Dict[blockName][regName]:
                            if regName in sheet_v1_Dict[blockName] and element !='Index' and element in sheet_v1_Dict[blockName][regName] and sheet_v2_Dict[blockName][regName][element] != sheet_v1_Dict[blockName][regName][element]:
                                flag=2

                    if flag ==1 :
                        output_changes.append(['Added']+[sheet]+sheet_v2.row_values(sheet_v2_Dict[blockName][regName]['Index']))
                    elif flag==2:
                        output_changes.append(['Changed From']+[sheet]+sheet_v1.row_values(sheet_v1_Dict[blockName][regName]['Index']))
                        output_changes.append(['Changed To']+[sheet]+sheet_v2.row_values(sheet_v2_Dict[blockName][regName]['Index']))
            for blockName in sheet_v1_Dict:
                for regName in sheet_v1_Dict[blockName]:
                    flag=0
                    if  (blockName not in sheet_v2_Dict) or (regName not in sheet_v2_Dict[blockName]):
                        flag=1
                    if flag !=0 :
                        output_changes.append(['Removed']+[sheet]+sheet_v1.row_values(sheet_v1_Dict[blockName][regName]['Index']))
            output_changes.append([''])
        output.save(filename2)

    def helper_new(self, swc_input):
        import webbrowser
        import operator
        self.PopulateAutoDict()
        swc_list  = []
        sorted_block_list = []
        #sorted_block_list = sorted(self.autoDict.keys(), key=lambda obj: obj, reverse=False)
        sorted_block_list = sorted(self.autoDict.keys(), key = lambda x: (self.autoDict[x]['DCB Binary Sub-Block']))
        print(sorted_block_list)
        for block in sorted_block_list:
            if self.autoDict[block]['Link'] not in swc_list:
                swc_list.append(self.autoDict[block]['Link'])

        if swc_input == '0':
            for i in range(len(swc_list)):
                webbrowser.open(swc_list[i])
        if swc_input != '0' and swc_input != '9':
            swc_value = int(swc_input) - 1
            webbrowser.open(swc_list[swc_value])

def ui_generate_dcb_binaries():
    # Make files writable
    for item in os.listdir(DSF.globals["release"]):
        if item.endswith(".bin"):
            os.chmod(DSF.globals["release"] + item, stat.S_IRWXU | stat.S_IRWXG | stat.S_IRWXO)
        
    if os.path.isfile(DSF.globals["target"] + "\sve\VVDRV_ddr_dcb.c"):
        os.chmod(DSF.globals["target"] + "\sve\VVDRV_ddr_dcb.c", stat.S_IRWXU | stat.S_IRWXG | stat.S_IRWXO)
        
    # Generate DCB binaries
    for filename in glob.glob(DSF.globals["target"] + platform_name + '*.xlsm'):
        dsf = status_message_and_time("Create DSF object", DSF, filename)
        msm_version = filename[-11:-10]
        silicon = filename[-9:-5]
        if silicon == 'RUMI':
            target_silicon = 0
        if silicon == 'ASIC':
            target_silicon = 1
        status_message_and_time("Generate DCB binary", dsf.GenerateDcbBinary, platform_id + '_0' + str(msm_version) + '00_' + str(target_silicon) + '_dcb.bin')
        
    # Generate DCB C file
    if os.path.isfile(DSF.globals["target"] + "\sve\gen_jtagless_dcb_c_file.py"):
        sys.path.append(DSF.globals["target"])
        import sve.gen_jtagless_dcb_c_file
    
def ui_modify_dcb_binaries():
    # Make files writable
    for item in os.listdir(DSF.globals["release"]):
        if item.endswith(".bin"):
            os.chmod(DSF.globals["release"] + item, stat.S_IRWXU | stat.S_IRWXG | stat.S_IRWXO)
        
    # Modify DCB binaries
    for filename in glob.glob(DSF.globals["external"] + platform_name + '*.xlsx'):
        dsf = status_message_and_time("Create DSF object", DSF, filename)
        msm_version = filename[-11:-10]
        silicon = filename[-9:-5]
        if silicon == 'RUMI':
            target_silicon = 0
        if silicon == 'ASIC':
            target_silicon = 1
        status_message_and_time("Update DCB binary from external DSF XL file", dsf.CustomerDcbBinary, platform_id + '_0' + str(msm_version) + '00_' + str(target_silicon) + '_dcb.bin')
        
    # Generate DCB C file
    if os.path.isfile(DSF.globals["target"] + "\sve\gen_jtagless_dcb_c_file.py"):
        sys.path.append(DSF.globals["target"])
        import sve.gen_jtagless_dcb_c_file

def ui_generate_external_dsf_xl_files():
    # Make files writable
    for item in os.listdir(DSF.globals["external"]):
        if item.endswith(".xlsx"):
            os.chmod(DSF.globals["external"] + item, stat.S_IRWXU | stat.S_IRWXG | stat.S_IRWXO)
        
    # Generate external DSF XL files
    for filename in glob.glob(DSF.globals["target"] + platform_name + '*.xlsm'):
        dsf = status_message_and_time("Create DSF object", DSF, filename)
        msm_version = filename[-11:-10]
        silicon = filename[-9:-5]
        if silicon == 'RUMI':
            target_silicon = 0
        if silicon == 'ASIC':
            target_silicon = 1
        status_message_and_time("Generate external DSF XL file", dsf.GenerateExternalDsfXl)
        status_message_and_time("Generate DCB binary", dsf.GenerateDcbBinary, platform_id + '_0' + str(msm_version) + '00_' + str(target_silicon) + '_dcb.bin')
        
    ui_modify_dcb_binaries()
    
def ui_generate_configmon():
    print("V1, V2, V3")
    msm_version = input("Enter msm version from above:")
    print("RUMI, ASIC")
    target_silicon = input("Enter target silicon from the above:")
    for filename in glob.glob(DSF.globals["target"] + platform_name + '*.xlsm'):
        if platform_name in filename:
            if msm_version in filename:
                if target_silicon in filename:
                    dsf = status_message_and_time("Create DSF object", DSF, filename)
                    status_message_and_time("Populate Run Configmon", dsf.RunConfigmon)

def ui_update_swc_sheet():
    print("V1, V2, V3")
    msm_version = input("Enter msm version from above:")
    print("RUMI, ASIC")
    target_silicon = input("Enter target silicon from the above:")
    for filename in glob.glob(DSF.globals["target"] + platform_name + '*.xlsm'):
        if platform_name in filename:
            if msm_version in filename:
                if target_silicon in filename:
                    dsf = status_message_and_time("Create DSF object", DSF, filename)
                    import shutil
                    swc_import = input("Download SWC XL file from sharepoint (Y/N):")
                    while True:
                        if swc_import == 'y' or swc_import == 'Y':
                            print("DSF XL file to import to from below:")
                            print("[0] - All SWCs")
                            print("[1] - ddr_phy_3.2.xlsm")
                            print("[2] - ddr_cc_3.2.xlsm")
                            print("[3] - DDRSS_Napali_V1.0_Swsettings.xlsx")
                            print("[4] - LLCC_Napali_v1.0_SWSettings.xlsx")
                            print("[5] - CABO_SWSettings_P3_dev.xlsx")
                            print("[6] - MCCC_SWSettings_Napali1.0.xlsx")
                            print("[7] - mem_noc_SWSettings.xlsx")
                            print("[8] - SHRM_Napali_V1_SWSettings.xlsx")
                            print("[9] - Use files from SWC folder")

                            swc_input = input("Use the above options to download SWCs:")
                            dsf.helper_new(swc_input)

                            if swc_input == '9' or swc_input == '0':
                                break
                        if swc_import == 'N' or swc_import == 'n':
                            break
                    if swc_import == 'y' or swc_import == 'Y':
                        sve = input("Press enter key to proceed when SWC download has completed.  You may need to give your browser permission to download the file.")
                    source_dir = "C:\\Users\\" + os.getlogin() + "\\Downloads\\"
                    dest_dir = DSF.globals["target"] +"\\swc"
                    if not os.path.exists(DSF.globals["target"] +"\\swc"):
                        os.makedirs(DSF.globals["target"] +"\\swc")
                    if swc_import == 'y' or swc_import == 'Y':
                        files_xlsx = glob.iglob(os.path.join(source_dir, "*.xlsx"))
                        files_xlsm = glob.iglob(os.path.join(source_dir, "*.xlsm"))
                        for file1 in files_xlsx:
                            if os.path.isfile(file1):
                                shutil.copy2(file1, dest_dir)
                        for file1 in files_xlsm:
                            if os.path.isfile(file1):
                                shutil.copy2(file1, dest_dir)
                    status_message_and_time("Update SWCs sheet", dsf.UpdateSwcSheet)

def ui_update_training_sheet():
    print("V1, V2, V3")
    msm_version = input("Enter msm version from above:")
    print("RUMI, ASIC")
    target_silicon = input("Enter target silicon from the above:")
    for filename in glob.glob(DSF.globals["target"] + platform_name + '*.xlsm'):
        if platform_name in filename:
            if msm_version in filename:
                if target_silicon in filename:
                    dsf = status_message_and_time("Create DSF object", DSF, filename)
                    status_message_and_time("Update TRAINING sheet", dsf.UpdateTrainingsheet)
                    status_message_and_time("Update DCC sheet", dsf.UpdateDCCsheet)
                    dsf.RunSwcMacros()

def ui_update_rel_sheet():
    print("V1, V2, V3")
    msm_version = input("Enter msm version from above:")
    print("RUMI, ASIC")
    target_silicon = input("Enter target silicon from the above:")
    for filename in glob.glob(DSF.globals["target"] + platform_name + '*.xlsm'):
        if platform_name in filename:
            if msm_version in filename:
                if target_silicon in filename:
                    dsf = status_message_and_time("Create DSF object", DSF, filename)
                    status_message_and_time("Update REL sheet", dsf.UpdateRelativeAddresses)
                    
def ui():
    sys.path.append('.')
    if DSF.globals["PERFORCE_PATHS"]:
        print("[d] = Generate External DSF XL Files & Binaries")
    print("[i] = Generate Binaries using Internal DSF XL Files")
    print("[e] = Modify Binaries using External DSF XL Files")
    if DSF.globals["PERFORCE_PATHS"]:
        print("[c] = Run ConfigMon")
        print("[s] = Import from Software Calculator")
        print("[t] = Update TRAINING and DCC sheet")
        print("[r] = Update REL sheet")
    value = input("Select an option by typing the letter next to the option and pressing Enter: ")

    if value == 'd':
        ui_generate_external_dsf_xl_files()
    elif value == 'i':
        ui_generate_dcb_binaries()
    elif value == 'e':
        ui_modify_dcb_binaries()
    elif value == 'c':
        ui_generate_configmon()
    elif value == 's':
        ui_update_swc_sheet()
    elif value == 't':
        ui_update_training_sheet()
    elif value == 'r':
        ui_update_rel_sheet()
    else:
        print("[ERROR] Invalid input")

##
# Use DSF class
##
if __name__ == "__main__":
    target_folder = sys.argv[1]        #SDM845_830
    platform_name = sys.argv[2]        #845
    platform_code_name = sys.argv[3]   #Napali
    platform_id = sys.argv[4]          #6000
    
    DSF.create_globals(target_folder)
    
    if glob.glob(DSF.globals["target"] + platform_name + '*.xlsm'):
        ui()
    else:
        ui_modify_dcb_binaries()
